import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import os as OS
from openpyxl.comments import Comment
from datetime import date
from datetime import datetime


load_dotenv()


class ExcelInventory:
    def __init__(self, file_excel_nuevaplaneacion, file_excel_controlalmacen) -> None:
        self.file_excel_nuevaplaneacion = file_excel_nuevaplaneacion
        self.file_excel_controlalmacen = file_excel_controlalmacen

    @property
    def excelNuevaPlaneacion(self):
        return pd.read_excel(
            self.file_excel_nuevaplaneacion["path"],
            sheet_name=self.file_excel_nuevaplaneacion["sheet"],
        )

    @property
    def excelControlAlmacen(self):
        return pd.read_excel(
            self.file_excel_controlalmacen["path"],
            sheet_name=self.file_excel_controlalmacen["sheet"],
            header=0,
        )

    @property
    def toExcelEntradas(self):
        return pd.read_excel(
            self.file_excel_controlalmacen["path"],
            sheet_name="ENTRADAS",
        )

    @property
    def toExcelSalidas(self):
        return pd.read_excel(
            self.file_excel_controlalmacen["path"],
            sheet_name="SALIDAS",
        )

    def validateInventory(self, excel_dataframe):
        dataframe_validate = excel_dataframe.copy()

        dataframe_whit_entradas = dataframe_validate[
            dataframe_validate["FECHA DE REGISTRO"] == "2024-07-25"
        ]
        columns_to_select = ["FECHA DE REGISTRO", "SKU", "CANTIDAD"]
        return dataframe_whit_entradas[columns_to_select]

    def inventoryDiscrepancy(self):
        entradas = self.validateInventory(self.toExcelEntradas).copy()
        salidas = self.validateInventory(self.toExcelSalidas).copy()
        nuevaPlaneacion = self.excelNuevaPlaneacion.copy()
        nuevaPlaneacion = nuevaPlaneacion[["ID PRODUCTO ", "INVENTARIO FISICO "]]
        sku = self.parse_inventory().copy()

        nuevaPlaneacion.columns = nuevaPlaneacion.columns.str.strip()
        sku.columns = sku.columns.str.strip()
        entradas.columns = entradas.columns.str.strip()
        salidas.columns = salidas.columns.str.strip()

        entradas_sum = entradas.groupby("SKU")["CANTIDAD"].sum().reset_index()
        salidas_sum = salidas.groupby("SKU")["CANTIDAD"].sum().reset_index()

        sku_combined = sku.merge(entradas_sum, on="SKU", how="left").merge(
            salidas_sum, on="SKU", how="left", suffixes=("_ENTRADAS", "_SALIDAS")
        )

        sku_combined["CANTIDAD_ENTRADAS"].fillna(0, inplace=True)
        sku_combined["CANTIDAD_SALIDAS"].fillna(0, inplace=True)

        sku_combined["INVENTARIO FINAL (TON)"] = (
            sku_combined["INVENTARIO (TON)"]
            - sku_combined["CANTIDAD_ENTRADAS"]
            + sku_combined["CANTIDAD_SALIDAS"]
        )
        sku_combined.rename(
            columns={
                "SKU": "ID PRODUCTO",
                "INVENTARIO FINAL (TON)": "INVENTARIO FINAL",
            },
            inplace=True,
        )

        sku_combined.rename(
            columns={
                "SKU": "ID PRODUCTO",
                "INVENTARIO FINAL (TON)": "INVENTARIO FINAL",
            },
            inplace=True,
        )
        comparacion = sku_combined.merge(nuevaPlaneacion, on="ID PRODUCTO", how="outer")

        comparacion["DIFERENCIA"] = (
            comparacion["INVENTARIO FINAL"] - comparacion["INVENTARIO FISICO"]
        )

        alertas = comparacion[comparacion["DIFERENCIA"] != 0]

        print("Comparación de Inventarios:")
        print(comparacion)
        print("\nAlertas de Diferencias:")
        print(
            alertas[
                ["ID PRODUCTO", "INVENTARIO FINAL", "INVENTARIO FISICO", "DIFERENCIA"]
            ]
        )

        return f'se encuentra discrepancia en los siguientes productos \n{alertas[alertas["DIFERENCIA"] >= 1]}'

    def parse_inventory(self):
        sku_inventary = self.excelControlAlmacen[["SKU", "INVENTARIO (TON)"]]

        sku_inventary.loc[:, "INVENTARIO (TON)"] = sku_inventary[
            "INVENTARIO (TON)"
        ].astype(int)

        return sku_inventary

    def join_excel(self, excel1, excel2):
        mapeo_inventario = excel2.set_index("SKU")["INVENTARIO (TON)"].to_dict()

        excel1["INVENTARIO FISICO "] = (
            excel1["ID PRODUCTO "]
            .map(mapeo_inventario)
            .fillna(excel1["INVENTARIO FISICO "])
        )
        return excel1

    def save_updated_excel(self, updated_excel):
        wb = load_workbook(self.file_excel_nuevaplaneacion["path"])
        ws = wb[self.file_excel_nuevaplaneacion["sheet"]]
        comments = {
            cell.coordinate: cell.comment
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
            for cell in row
            if cell.comment
        }

        ws.delete_rows(2, ws.max_row)

        for r_idx, row in enumerate(
            dataframe_to_rows(updated_excel, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if cell.coordinate in comments:
                    cell.comment = comments[cell.coordinate]

        wb.save(self.file_excel_nuevaplaneacion["path"])

    def current_date_format(self, date_now):
        day = date_now.day - 1
        month = date_now.month
        year = date_now.year

        message = "{:04d}-{:02d}-{:02d}".format(year, month, day)
        return str(message)

    def bl_excel(self):
        regex_pattern = r"^(?=.*[A-Za-z])|^.{4,}$"

        despreciated_value_NA = self.toExcelEntradas[
            (self.toExcelEntradas["OC / BL"] != "N/A")
            & (self.toExcelEntradas["OC / BL"].str.strip() != "NaN")
            & (self.toExcelEntradas["OC / BL"] != "-")
            & (self.toExcelEntradas["OC / BL"].str.strip().str.contains(regex_pattern))
        ]
        despreciated_value = despreciated_value_NA[
            despreciated_value_NA["PROVEEDOR"] != "FORJACHISA"
        ]
        last_entries = despreciated_value[
            despreciated_value["FECHA DE REGISTRO"] == "2024-07-25"
        ]
        dataframe_excelNuevaPlaneacion = self.excelNuevaPlaneacion
        inventory_last = last_entries.set_index("SKU")["CANTIDAD"].to_dict()
        inventory_map = dataframe_excelNuevaPlaneacion["ID PRODUCTO "].map(
            inventory_last
        )

        # dataframe_excelNuevaPlaneacion["INVENTARIO EN TRANSITO "] = (
        #     dataframe_excelNuevaPlaneacion["INVENTARIO EN TRANSITO "]
        #     - inventory_map.fillna(0)
        # )

        workbook = load_workbook(self.file_excel_nuevaplaneacion["path"])
        sheet = workbook["INVENTARIO ACTUAL "]

        # formula_column = "D"
        # for row in sheet.iter_rows(
        #     min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
        # ):
        #     for cell in row:
        #         if (
        #             cell.value
        #             and isinstance(cell.value, str)
        #             and cell.value.startswith("=")
        #         ):
        #             formula_cell = f"{formula_column}{cell.row}"
        #             sheet[formula_cell] = cell.value

        # for producto_id, amount_to_subtract in inventory_last.items():
        #     amount_to_subtract = float(amount_to_subtract)
        #     producto_id = producto_id.lower().strip().replace(" ", "")
        #     producto_fila = None

        #     for index, row in dataframe_excelNuevaPlaneacion.iterrows():
        #         if row["ID PRODUCTO "].lower().strip().replace(" ", "") == producto_id:
        #             producto_fila = index + 2
        #             break

        #     if producto_fila:
        #         cell = sheet[f"D{producto_fila}"]

        #         if (
        #             cell.value
        #             and isinstance(cell.value, str)
        #             and cell.value.startswith("=")
        #         ):
        #             original_formula = cell.value.split("=")[1]
        #             modified_formula = f"=({original_formula}) - {amount_to_subtract}"
        #             cell.value = modified_formula
        #             print(
        #                 f"Fórmula modificada en la celda D{producto_fila}: {modified_formula}"
        #             )
        #         else:
        #             print(f"No se encontró una fórmula en la celda D{producto_fila}")
        #     else:
        #         print(f"Producto con ID {producto_id} no encontrado.")

        # workbook.save(self.file_excel_nuevaplaneacion["path"])

        comments = {
            cell.coordinate: cell.comment
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row)
            for cell in row
            if cell.comment
        }

        sheet.delete_rows(2, sheet.max_row)

        for r_idx, row in enumerate(
            dataframe_to_rows(dataframe_excelNuevaPlaneacion, index=False, header=True),
            1,
        ):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                if cell.coordinate in comments:
                    cell.comment = comments[cell.coordinate]

        workbook.save(self.file_excel_nuevaplaneacion["path"])

        # Eliminar comentarios específicos
        bl = last_entries.set_index("SKU")["OC / BL"].to_dict()

        for producto_id, comment_to_remove in bl.items():
            comment_to_remove = str(comment_to_remove)
            producto_id = producto_id.lower().strip().replace(" ", "")
            producto_fila = None
            for index, row in dataframe_excelNuevaPlaneacion.iterrows():
                if row["ID PRODUCTO "].lower().strip().replace(" ", "") == producto_id:
                    producto_fila = index + 2
                    break

            if producto_fila:
                cell = sheet[f"D{producto_fila}"]

                if cell.comment:
                    existing_comment = str(cell.comment.text)
                    if comment_to_remove in existing_comment:
                        new_comment_text = existing_comment.replace(
                            comment_to_remove, ""
                        ).strip()
                        if new_comment_text:
                            cell.comment = Comment(
                                new_comment_text, cell.comment.author
                            )
                        else:
                            cell.comment = None
                        print(
                            f"Comentario '{comment_to_remove}' eliminado de la celda D{producto_fila}"
                        )
                    else:
                        print(
                            f"El comentario '{comment_to_remove}' no se encontró en la celda E{producto_fila}"
                        )
                else:
                    print(f"No hay comentario existente en la celda E{producto_fila}")
            else:
                print(f"Producto con ID {producto_id} no encontrado.")

        workbook.save(self.file_excel_nuevaplaneacion["path"])

        column_index = dataframe_excelNuevaPlaneacion.columns.get_loc(
            "ORDENES COLOCADAS "
        )
        return dataframe_excelNuevaPlaneacion.iloc[:, : column_index + 1]

    @property
    def process(self):
        nueva_planeacion = self.excelNuevaPlaneacion
        alerta = self.inventoryDiscrepancy()
        print(alerta)
        parsed_inventory = self.parse_inventory()
        updated_excel = self.join_excel(nueva_planeacion, parsed_inventory)
        update_finalexcel = self.bl_excel()
        self.save_updated_excel(updated_excel)

        return update_finalexcel


if __name__ == "__main__":
    nueva_planeacion_path = OS.getenv("NUEVA_PLANEACION_PATH")
    control_almacen_path = OS.getenv("CONTROL_ALMACEN_PATH")
    print(nueva_planeacion_path)

    inventory = ExcelInventory(
        {"path": nueva_planeacion_path, "sheet": "INVENTARIO ACTUAL "},
        {"path": control_almacen_path, "sheet": "SKU - COMPRAS"},
    )
    update_excel = inventory.process

    print(update_excel)
