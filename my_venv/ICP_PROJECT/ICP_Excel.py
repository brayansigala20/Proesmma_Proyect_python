import pandas as pd
import os
from openpyxl import Workbook
import datetime

excel = pd.read_excel(
    "C:\\Users\\braya\\Downloads\\ICP\\MASTER ROBOT ICP 2024.xlsx"
)
header_datetime = datetime.datetime(2024, 5, 1, 0, 0)

columnData = {}
if header_datetime in excel.columns:
    columnData = excel[header_datetime]

else:
    print(f'No se encontró la columna con el encabezado {header_datetime}')

print(columnData)
string = excel["CLIENTE"].to_dict()
wb = Workbook()

ws = wb.active
ws.title = "Lista de Nombres"

ws["A1"] = "Nombre"
ws["B1"] = "Folio"

for idx, item in enumerate(string.values(), start=2):
    text = item
    ws.cell(row=idx, column=1, value=text)

for idx, item in enumerate(columnData, start=2):
    text = item
    ws.cell(row=idx, column=2, value=text)


wb.save("C:\\Users\\braya\\Downloads\\ICP\\Lista_Nombres.xlsx")
